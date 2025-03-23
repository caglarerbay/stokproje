# inventory/utils.py
import openpyxl
import io
from openpyxl.styles import Alignment
from io import BytesIO
from pyfcm import FCMNotification
from django.conf import settings
from .models import Device
from firebase_admin import messaging
from django.conf import settings

def create_excel_report(critical_products):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kritik Stok Raporu"

    # Başlık satırı
    headers = ["Parça Kodu", "Ürün İsmi", "Mevcut Adet", "Min Limit", "Sipariş Durumu"]
    ws.append(headers)

    # Sütun başlıklarını ortala
    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
        for cell in col:
            cell.alignment = Alignment(horizontal="center")

    # Verileri ekle
    for product in critical_products:
        durum = "Siparişi çekilmiştir, tedariği bekleniyor." if product.order_placed else "Kritik stok, sipariş henüz çekilmemiş."
        row = [product.part_code, product.name, product.quantity, product.min_limit, durum]
        ws.append(row)

    # Excel dosyasını bayt dizisi olarak döndürmek için
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_excel_for_products(products):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ana Stok"

    headers = ["part_code", "name", "quantity", "min_limit", "order_placed"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    row_idx = 2
    for p in products:
        ws.cell(row=row_idx, column=1, value=p.part_code)
        ws.cell(row=row_idx, column=2, value=p.name)
        ws.cell(row=row_idx, column=3, value=p.quantity)
        ws.cell(row=row_idx, column=4, value=p.min_limit)
        ws.cell(row=row_idx, column=5, value=str(p.order_placed))
        row_idx += 1

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file

def create_excel_for_single_user(user, user_stocks):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{user.username} Stok"

    headers = ["part_code", "product_name", "quantity"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    row_idx = 2
    for us in user_stocks:
        ws.cell(row=row_idx, column=1, value=us.product.part_code)
        ws.cell(row=row_idx, column=2, value=us.product.name)
        ws.cell(row=row_idx, column=3, value=us.quantity)
        row_idx += 1

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file




def send_push_notification(message_title, message_body):
    # Şimdilik sabit token listesi ile test edelim.
    # Gerçek cihaz token'larını mobil uygulama FCM entegrasyonuyla alacaksın.
    registration_ids = ['YOUR_TEST_DEVICE_TOKEN']  # Bu değeri test token ile değiştir.

    if not registration_ids:
        return None

    # Multicast mesaj oluşturuyoruz.
    message = messaging.MulticastMessage(
        notification=messaging.Notification(
            title=message_title,
            body=message_body,
        ),
        tokens=registration_ids,
    )

    response = messaging.send_multicast(message)
    return response